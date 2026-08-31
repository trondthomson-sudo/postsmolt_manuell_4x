"""
excel_export.py
------------------
Skriver ukekalenderen (scheduler.build_schedule sin wide DataFrame) til en
fargekodet .xlsx som følger samme oppsett som illustrasjonen: bloc per tank,
fet status, tallformatert biomasse/WFE.
"""
from __future__ import annotations
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILL_HEADER = PatternFill("solid", fgColor="1F3B57")
FILL_TANK1 = PatternFill("solid", fgColor="E3EFE1")
FILL_TANK2 = PatternFill("solid", fgColor="DCEAF5")
FILL_TANK3 = PatternFill("solid", fgColor="FBE9DD")
FILL_MISC = PatternFill("solid", fgColor="FBF3D9")
FILL_CLEAN = PatternFill("solid", fgColor="C9C9C9")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
FONT_LABEL = Font(bold=True, size=10)
FONT_BOLD = Font(bold=True, size=10)
FONT_NORMAL = Font(size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROW_FILLS = {
    "Tank 1 - kohort": FILL_TANK1, "Tank 1 - status": FILL_TANK1, "Tank 1 - biomasse (t)": FILL_TANK1,
    "Tank 1 - tetthet (kg/m3)": FILL_TANK1, "Tank 1 - vekt (g)": FILL_TANK1,
    "Tank 2 - kohort": FILL_TANK2, "Tank 2 - status": FILL_TANK2, "Tank 2 - biomasse (t)": FILL_TANK2,
    "Tank 2 - tetthet (kg/m3)": FILL_TANK2, "Tank 2 - vekt (g)": FILL_TANK2,
    "Tank 3 - kohort": FILL_TANK3, "Tank 3 - status": FILL_TANK3, "Tank 3 - biomasse (t)": FILL_TANK3,
    "Tank 3 - tetthet (kg/m3)": FILL_TANK3, "Tank 3 - vekt (g)": FILL_TANK3,
    "Overforing": FILL_MISC, "Levering": FILL_MISC, "Levert WFE (t)": FILL_MISC, "Akkumulert i aret (t)": FILL_MISC,
}


def write_excel(df: pd.DataFrame, path: str, max_weeks: int | None = None, title: str = "Postsmolt-kalender"):
    cols = list(df.columns)[:max_weeks] if max_weeks else list(df.columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kalender"

    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=13)
    ws.freeze_panes = "C4"

    header_row = 3
    ws.cell(row=header_row, column=1, value="Felt").fill = FILL_HEADER
    ws.cell(row=header_row, column=1).font = FONT_HEADER
    for j, col in enumerate(cols, start=2):
        c = ws.cell(row=header_row, column=j, value=col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    for i, (label, row) in enumerate(df.iterrows(), start=header_row + 1):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = FONT_LABEL
        fill = ROW_FILLS.get(label)
        if fill:
            lc.fill = fill
        is_status_row = label.endswith("status")
        is_number_row = ("biomasse" in label or "WFE" in label or "Akkumulert" in label
                          or "tetthet" in label or label.endswith("vekt (g)"))
        for j, col in enumerate(cols, start=2):
            val = row[col]
            cell = ws.cell(row=i, column=j)
            if is_number_row:
                is_weight_row = label.endswith("vekt (g)")
                # "" (eldre tomme celler) OG None/NaN (dagens tomme
                # vekt-celler, se scheduler_manuell.py) betyr begge "ingen
                # verdi denne uken" - float(None) krasjer, sa begge ma
                # fanges opp her for tomme rader.
                if val == "" or pd.isna(val):
                    cell.value = None if is_weight_row else 0.0
                else:
                    cell.value = float(val)
                cell.number_format = "#,##0.0"
            else:
                cell.value = val
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if is_status_row and val == "Rengjoring":
                cell.fill = FILL_CLEAN
                cell.font = FONT_BOLD
            elif is_status_row and val == "Vekst":
                cell.font = FONT_BOLD
            elif label.endswith("kohort") and val:
                cell.font = FONT_BOLD

    ws.column_dimensions["A"].width = 24
    for j in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 9.5

    wb.save(path)
